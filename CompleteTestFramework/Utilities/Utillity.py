from openpyxl.reader.excel import load_workbook

from Base.test_Basedriver import BaseDriver


class Utils(BaseDriver):

    def __init__(self, driver):
        super().__init__(driver)
        self.driver = driver

    def window_tabs(self):
        current_handle = self.driver.current_window_handle
        self.driver.switch_to.window(current_handle)

    def switch_to_windowtab(self,current_handle):
        self.driver.switch_to.window(current_handle)

    def window_list(self,index):
        handle_list = self.driver.window_handles
        self.driver.switch_to.window(handle_list[index])


    def switch_to_window(self,handle_index):
        self.driver.switch_to.window(handle_index)

    def read_from_excel(filename,sheet):
        wl = load_workbook(filename)
        sh = wl.active

        data_list = []
        row_ct =  sh.max_row
        col_ct =  sh.max_column

        for i in range(2, row_ct+1):
            row = []
            for j in range(1,col_ct+1):
                row.append(sh.cell(row=i, column=j).value)
            data_list.append(row)
        return data_list


